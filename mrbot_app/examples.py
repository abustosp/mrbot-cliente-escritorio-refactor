import os
import sys
import pathlib
from typing import Dict, Tuple, Union

# Ajustar sys.path si se ejecuta directamente (python mrbot_app/examples.py)
if __package__ is None or __package__ == "":
    sys.path.append(str(pathlib.Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mrbot_app.constants import EXAMPLE_DIR, ACCENT, FG
from mrbot_app.formatos import aplicar_formato_encabezado, autoajustar_columnas, agregar_filtros


def ensure_example_excels() -> Dict[str, str]:
    """
    Crea archivos Excel de ejemplo para cada endpoint si no existen.
    Retorna un dict con el nombre corto -> ruta.
    """
    os.makedirs(EXAMPLE_DIR, exist_ok=True)

    # DataFrames for single-sheet Excels
    examples: Dict[str, pd.DataFrame] = {
        "mis_comprobantes.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_inicio_sesion": "20123456789",
                    "nombre_representado": "Empresa Demo SA",
                    "cuit_representado": "20987654321",
                    "contrasena": "clave_demo",
                    "descarga_emitidos": "SI",
                    "descarga_recibidos": "SI",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_emitidos": "/tmp/emitidos",
                    "nombre_emitidos": "emitidos-demo",
                    "ubicacion_recibidos": "/tmp/recibidos",
                    "nombre_recibidos": "recibidos-demo",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_inicio_sesion": "20111111111",
                    "nombre_representado": "Ejemplo NO",
                    "cuit_representado": "20999999999",
                    "contrasena": "clave_no",
                    "descarga_emitidos": "NO",
                    "descarga_recibidos": "NO",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_emitidos": "/tmp/emitidos",
                    "nombre_emitidos": "emitidos-no",
                    "ubicacion_recibidos": "/tmp/recibidos",
                    "nombre_recibidos": "recibidos-no",
                    "retry": "0",
                },
            ]
        ),
        "rcel.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "nombre_rcel": "Empresa Demo SA",
                    "representado_cuit": "20987654321",
                    "clave": "clave_demo",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/RCEL/20987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "nombre_rcel": "Ejemplo NO",
                    "representado_cuit": "20999999999",
                    "clave": "clave_no",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/RCEL/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "hacienda.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "denominacion": "Empresa Demo Hacienda SA",
                    "representado_cuit": "20987654321",
                    "clave": "clave_demo",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/Hacienda/20987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "denominacion": "Ejemplo NO Hacienda",
                    "representado_cuit": "20999999999",
                    "clave": "clave_no",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/Hacienda/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "liquidacion_granos.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "clave": "clave_demo",
                    "denominacion": "Empresa Demo Granos SA",
                    "cuit_representado": "20987654321",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/Liquidacion_Granos/20987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave": "clave_no",
                    "denominacion": "Ejemplo NO Granos",
                    "cuit_representado": "",
                    "desde": "01/01/2024",
                    "hasta": "31/12/2024",
                    "ubicacion_descarga": "./descargas/Liquidacion_Granos/20111111111",
                    "retry": "0",
                },
            ]
        ),
        "sct.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_login": "20123456789",
                    "cuit_representado": "20987654321",
                    "clave": "clave_demo",
                    "deuda": "SI",
                    "vencimientos": "SI",
                    "presentacion_ddjj": "SI",
                    "excel": "SI",
                    "csv": "SI",
                    "pdf": "NO",
                    "ubicacion_deuda": "./Descargas",
                    "nombre_deuda": "deuda-demo",
                    "ubicacion_vencimientos": "./Descargas",
                    "nombre_vencimientos": "vencimientos-demo",
                    "ubicacion_ddjj": "./Descargas",
                    "nombre_ddjj": "ddjj-demo",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_login": "20111111111",
                    "cuit_representado": "20999999999",
                    "clave": "clave_no",
                    "deuda": "NO",
                    "vencimientos": "NO",
                    "presentacion_ddjj": "NO",
                    "excel": "NO",
                    "csv": "NO",
                    "pdf": "NO",
                    "ubicacion_deuda": "./Descargas",
                    "nombre_deuda": "deuda-no",
                    "ubicacion_vencimientos": "./Descargas",
                    "nombre_vencimientos": "vencimientos-no",
                    "ubicacion_ddjj": "./Descargas",
                    "nombre_ddjj": "ddjj-no",
                    "retry": "0",
                },
            ]
        ),
        "ccma.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "clave_representante": "clave_demo",
                    "cuit_representado": "20987654321",
                    "movimientos": "SI",
                    "pdf": "SI",
                    "ubicacion_descarga": "./descargas/CCMA/20987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave_representante": "clave_no",
                    "cuit_representado": "20999999999",
                    "movimientos": "NO",
                    "pdf": "NO",
                    "ubicacion_descarga": "./descargas/CCMA/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "mis_retenciones.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "clave_representante": "tu_clave_fiscal",
                    "cuit_representado": "30987654321",
                    "denominacion": "Empresa Ejemplo SA",
                    "desde": "01/11/2025",
                    "hasta": "30/11/2025",
                    "ubicacion_descarga": "./descargas/Mis_Retenciones/30987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave_representante": "clave_no",
                    "cuit_representado": "20999999999",
                    "denominacion": "Ejemplo NO",
                    "desde": "01/01/2024",
                    "hasta": "31/01/2024",
                    "ubicacion_descarga": "./descargas/Mis_Retenciones/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "sifere.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "27123456789",
                    "clave_representante": "tu_clave_fiscal",
                    "cuit_representado": "20987654321",
                    "periodo": "202401",
                    "representado_nombre": "Empresa Ejemplo SA",
                    "jurisdicciones": "todas",
                    "ubicacion_descarga": "./descargas/SIFERE/20987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave_representante": "clave_no",
                    "cuit_representado": "20999999999",
                    "periodo": "202312",
                    "representado_nombre": "Ejemplo NO",
                    "jurisdicciones": "901,902;903|904",
                    "ubicacion_descarga": "./descargas/SIFERE/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "declaracion_en_linea.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "clave_representante": "tu_clave_fiscal",
                    "cuit_representado": "30987654321",
                    "representado_nombre": "Empresa Ejemplo SA",
                    "periodo_desde": "202511",
                    "periodo_hasta": "202511",
                    "ubicacion_descarga": "./descargas/Declaracion_en_Linea/30987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave_representante": "clave_no",
                    "cuit_representado": "20999999999",
                    "representado_nombre": "Ejemplo NO",
                    "periodo_desde": "202401",
                    "periodo_hasta": "202412",
                    "ubicacion_descarga": "./descargas/Declaracion_en_Linea/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "mis_facilidades.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_login": "20123456789",
                    "clave": "tu_clave_fiscal",
                    "cuit_representado": "30987654321",
                    "denominacion": "Empresa Ejemplo SA",
                    "ubicacion_descarga": "./descargas/Mis_Facilidades/30987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_login": "20111111111",
                    "clave": "clave_no",
                    "cuit_representado": "20999999999",
                    "denominacion": "Ejemplo NO",
                    "ubicacion_descarga": "./descargas/Mis_Facilidades/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "pago_devoluciones.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_representante": "20123456789",
                    "clave_representante": "tu_clave_fiscal",
                    "cuit_representado": "30987654321",
                    "proxy_request": "NO",
                    "carga_minio": "SI",
                    "ubicacion_descarga": "./descargas/Pago_Devoluciones/30987654321",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_representante": "20111111111",
                    "clave_representante": "clave_no",
                    "cuit_representado": "20999999999",
                    "proxy_request": "NO",
                    "carga_minio": "SI",
                    "ubicacion_descarga": "./descargas/Pago_Devoluciones/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "aportes_en_linea.xlsx": pd.DataFrame(
            [
                {
                    "procesar": "SI",
                    "cuit_login": "20123456789",
                    "clave": "tu_clave_fiscal",
                    "cuit_representado": "20123456789",
                    "ubicacion_descarga": "./descargas/Aportes_en_Linea/20123456789",
                    "retry": "0",
                },
                {
                    "procesar": "NO",
                    "cuit_login": "20111111111",
                    "clave": "clave_no",
                    "cuit_representado": "20999999999",
                    "ubicacion_descarga": "./descargas/Aportes_en_Linea/20999999999",
                    "retry": "0",
                },
            ]
        ),
        "apocrifos.xlsx": pd.DataFrame(
            [
                {"cuit": "20333444555"},
                {"cuit": "27999888777"},
            ]
        ),
        "consulta_cuit.xlsx": pd.DataFrame([{"cuit": "20333444555"}, {"cuit": "20987654321"}]),
        "control_monotributistas.xlsx": pd.DataFrame(
            [
                {
                    "CUIT_Representante": "20123456789",
                    "Clave_representante": "clave_demo",
                    "CUIT_Representado": "20987654321",
                    "Denominacion_MC": "Empresa Demo MC",
                    "Denominacion_RCEL": "Empresa Demo RCEL",
                    "Descarga_MC": "SI",
                    "Descarga_MC_emitidos": "SI",
                    "Descarga_MC_recibidos": "NO",
                    "Desde_MC": "01/01/2024",
                    "Hasta_MC": "31/12/2024",
                    "Descarga_RCEL": "SI",
                    "Desde_RCEL": "01/01/2024",
                    "Hasta_RCEL": "31/12/2024",
                    "Ubicacion_Descarga_MC": "",
                    "Ubicacion_Descarga_RCEL": ""
                },
                {
                    "CUIT_Representante": "20111111111",
                    "Clave_representante": "clave_no",
                    "CUIT_Representado": "20999999999",
                    "Denominacion_MC": "Empresa NO MC",
                    "Denominacion_RCEL": "Empresa NO RCEL",
                    "Descarga_MC": "NO",
                    "Descarga_MC_emitidos": "NO",
                    "Descarga_MC_recibidos": "NO",
                    "Desde_MC": "01/01/2024",
                    "Hasta_MC": "31/12/2024",
                    "Descarga_RCEL": "NO",
                    "Desde_RCEL": "01/01/2024",
                    "Hasta_RCEL": "31/12/2024",
                    "Ubicacion_Descarga_MC": "",
                    "Ubicacion_Descarga_RCEL": ""
                },
            ]
        ),
    }

    paths: Dict[str, str] = {}

    # Process standard files
    for name, df in examples.items():
        path = os.path.join(EXAMPLE_DIR, name)
        paths[name] = path
        expected_cols = [c.strip().lower() for c in df.columns]
        should_write = not os.path.exists(path)
        if not should_write and name == "ccma.xlsx":
            try:
                current_cols = [c.strip().lower() for c in pd.read_excel(path, nrows=0).columns]
                if any(col not in current_cols for col in expected_cols):
                    should_write = True
            except Exception:
                should_write = True

        # Override for control_monotributistas to always ensure columns match if needed, but for now standard check is fine

        if should_write:
            try:
                df.to_excel(path, index=False)
                _format_excel(path)
            except Exception:
                pass
        else:
            _format_excel(path)

    # Process Categorias.xlsx (Multi-sheet)
    path_categorias = os.path.join(EXAMPLE_DIR, "Categorias.xlsx")
    paths["Categorias.xlsx"] = path_categorias
    if not os.path.exists(path_categorias):
        try:
            with pd.ExcelWriter(path_categorias, engine='openpyxl') as writer:
                # Sheet Categorias
                df_cat = pd.DataFrame({
                    "Categoria": ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"],
                    "Ingresos brutos": [2108288.01, 3133941.63, 4387518.23, 5449094.55, 6416528.72, 8020660.9, 9624793.05, 11916410.45, 13337213.22, 15285088.06, 16957968.71]
                })
                df_cat.to_excel(writer, sheet_name="Categorias", index=False)

                # Sheet Rango de Fechas
                # A2: Start Date, B2: End Date
                # We can create a dataframe that puts values in row 1 (since header is row 0)
                # Or just a dataframe with "Desde", "Hasta" and one row
                df_fechas = pd.DataFrame({
                    "Desde": ["01/01/2024"],
                    "Hasta": ["31/12/2024"]
                })
                # The code reads A2 and B2, which corresponds to the first data row if headers are present
                df_fechas.to_excel(writer, sheet_name="Rango de Fechas", index=False)

            _format_excel(path_categorias)
        except Exception as e:
            pass

    return paths


def _format_excel(path: str) -> None:
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            aplicar_formato_encabezado(ws)
            autoajustar_columnas(ws)
            # agregar_filtros(ws) # Apply only to some? Safe to apply to all generally
        wb.save(path)
    except Exception:
        pass


if __name__ == "__main__":
    rutas = ensure_example_excels()
    for k, v in rutas.items():
        print(f"{k}: {v}")
