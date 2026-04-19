import json
import os
import re
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from mrbot_app.formatos import aplicar_formato_encabezado, autoajustar_columnas, agregar_filtros
from mrbot_app.helpers import build_headers, ensure_trailing_slash, get_unique_filename, safe_post


def _log(message: str, log_fn: Optional[Callable[[str], None]] = None) -> None:
    if log_fn:
        log_fn(message)


def sanitize_identifier(value: str, fallback: str = "sin_cuit") -> str:
    cleaned = re.sub(r"[^0-9A-Za-z._-]", "_", (value or "").strip())
    cleaned = cleaned.strip("_")
    return cleaned or fallback


def parse_cuits_input(value: Any) -> List[str]:
    if value is None:
        return []

    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value).strip()
        if not text:
            return []
        raw_items = [part.strip() for part in re.split(r"[,;|\n\t ]+", text) if part.strip()]

    cuits: List[str] = []
    seen = set()
    for item in raw_items:
        cuit = re.sub(r"\D", "", str(item))
        if not cuit:
            continue
        if cuit in seen:
            continue
        seen.add(cuit)
        cuits.append(cuit)
    return cuits


def build_srt_payload(
    cuit_login: str,
    clave: str,
    cuits_consulta: Any,
    proxy_request: Optional[bool] = None,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "cuit_login": (cuit_login or "").strip(),
        "clave": clave or "",
        "cuits_consulta": parse_cuits_input(cuits_consulta),
    }
    if proxy_request is not None:
        payload["proxy_request"] = bool(proxy_request)
    return payload


def redact_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    safe = dict(payload)
    if "clave" in safe:
        safe["clave"] = "***"
    return safe


def consultar_srt_alicuotas(
    base_url: str,
    api_key: str,
    email: str,
    cuit_login: str,
    clave: str,
    cuits_consulta: Any,
    proxy_request: Optional[bool] = None,
    timeout_sec: Optional[int] = None,
) -> Dict[str, Any]:
    headers = build_headers(api_key, email)
    payload = build_srt_payload(cuit_login, clave, cuits_consulta, proxy_request)
    url = ensure_trailing_slash(base_url) + "api/v1/srt/alicuotas/consulta"
    response = safe_post(url, headers, payload, timeout_sec=timeout_sec)
    response["request_payload"] = payload
    response["request_payload_safe"] = redact_payload(payload)
    response["request_url"] = url
    return response


def _to_float(value: str) -> Optional[float]:
    text = str(value or "").strip()
    if not text:
        return None

    # Normalize decimal separators for robust parsing.
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _parse_alicuota_text(alicuota_text: str) -> tuple[Optional[float], Optional[float]]:
    text = str(alicuota_text or "")
    var_match = re.search(r"variable\s*:\s*([0-9.,-]+)%", text, flags=re.IGNORECASE)
    fija_match = re.search(r"suma\s*fija\s*:\s*\$\s*([0-9.,-]+)", text, flags=re.IGNORECASE)

    suma_variable = _to_float(var_match.group(1)) if var_match else None
    suma_fija = _to_float(fija_match.group(1)) if fija_match else None
    return suma_fija, suma_variable


def _parse_ciiu_text(ciiu_text: str) -> tuple[str, str]:
    text = str(ciiu_text or "").strip()
    if not text:
        return "", ""

    match = re.match(r"^(\d+)\s*-\s*(.+)$", text)
    if not match:
        return "", text

    return match.group(1).strip(), match.group(2).strip()


def _extract_ok_block_values(block: Any) -> tuple[str, str, Optional[float], Optional[float]]:
    if not isinstance(block, dict):
        return "", "", None, None

    rows = block.get("rows")
    if not isinstance(rows, list):
        return "", "", None, None

    ciiu_text = ""
    alicuota_text = ""
    for row in rows:
        if not isinstance(row, list) or len(row) < 2:
            continue

        key = str(row[0]).strip().lower()
        value = str(row[1]).strip()
        if "ciiu" in key:
            ciiu_text = value
        if "alicuota" in key or "alícuota" in key:
            alicuota_text = value

    ciiu_num, ciiu_desc = _parse_ciiu_text(ciiu_text)
    suma_fija, suma_variable = _parse_alicuota_text(alicuota_text)
    return ciiu_num, ciiu_desc, suma_fija, suma_variable


def normalize_srt_consulta_rows(consultas: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not isinstance(consultas, list):
        return rows

    for consulta in consultas:
        if not isinstance(consulta, dict):
            rows.append(
                {
                    "CUIT": "",
                    "afiliacion": "error de formato en consulta",
                    "CIUU (numero)": "",
                    "Descripción CIIU": "",
                    "suma fija": None,
                    "suma variable": None,
                }
            )
            continue

        cuit = str(consulta.get("cuit", "") or "").strip()
        status = str(consulta.get("status", "") or "").strip().upper()
        data = consulta.get("data")
        if status == "SIN_AFILIACION_VIGENTE":
            rows.append(
                {
                    "CUIT": cuit,
                    "afiliacion": "consultado no tiene afiliación vigente",
                    "CIUU (numero)": "",
                    "Descripción CIIU": "",
                    "suma fija": None,
                    "suma variable": None,
                }
            )
            continue

        if status != "OK":
            message = str(consulta.get("message", "") or "").strip() or status or "error de consulta"
            rows.append(
                {
                    "CUIT": cuit,
                    "afiliacion": message,
                    "CIUU (numero)": "",
                    "Descripción CIIU": "",
                    "suma fija": None,
                    "suma variable": None,
                }
            )
            continue

        if not isinstance(data, list) or not data:
            rows.append(
                {
                    "CUIT": cuit,
                    "afiliacion": "sin datos de alícuota",
                    "CIUU (numero)": "",
                    "Descripción CIIU": "",
                    "suma fija": None,
                    "suma variable": None,
                }
            )
            continue

        for block in data:
            ciiu_num, ciiu_desc, suma_fija, suma_variable = _extract_ok_block_values(block)
            rows.append(
                {
                    "CUIT": cuit,
                    "afiliacion": "",
                    "CIUU (numero)": ciiu_num,
                    "Descripción CIIU": ciiu_desc,
                    "suma fija": suma_fija,
                    "suma variable": suma_variable,
                }
            )

    return rows


def save_raw_response_json(raw_response: Any, file_path: str) -> str:
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    with open(file_path, "w", encoding="utf-8") as handle:
        json.dump(raw_response, handle, ensure_ascii=False, indent=2, default=str)
    return file_path


def save_consultas_json_by_cuit(consultas: Any, base_dir: str) -> List[str]:
    os.makedirs(base_dir, exist_ok=True)
    saved: List[str] = []

    if not isinstance(consultas, list):
        return saved

    for consulta in consultas:
        if not isinstance(consulta, dict):
            continue
        cuit = sanitize_identifier(str(consulta.get("cuit", "") or ""))
        cuit_dir = os.path.join(base_dir, cuit)
        os.makedirs(cuit_dir, exist_ok=True)
        base_name = f"srt_alicuotas_{cuit}.json"
        filename = get_unique_filename(cuit_dir, base_name)
        target = os.path.join(cuit_dir, filename)
        with open(target, "w", encoding="utf-8") as handle:
            json.dump(consulta, handle, ensure_ascii=False, indent=2, default=str)
        saved.append(target)

    return saved


def save_consolidated_excel(rows: List[Dict[str, Any]], output_dir: str, filename: Optional[str] = None) -> Optional[str]:
    if not rows:
        return None

    os.makedirs(output_dir, exist_ok=True)
    default_name = filename or f"srt_alicuotas_consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    unique_name = get_unique_filename(output_dir, default_name)
    output_path = os.path.join(output_dir, unique_name)

    df = pd.DataFrame(rows)
    if df.empty:
        return None

    ordered_columns = [
        "CUIT",
        "afiliacion",
        "CIUU (numero)",
        "Descripción CIIU",
        "suma fija",
        "suma variable",
    ]
    existing = [col for col in ordered_columns if col in df.columns]
    extras = [col for col in df.columns if col not in existing]
    df = df[existing + extras]

    if "suma variable" in df.columns:
        # API values come as percentage points (e.g. 4.800), convert to ratio for Excel percent formatting.
        df["suma variable"] = pd.to_numeric(df["suma variable"], errors="coerce") / 100.0

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Consolidado"

    suma_variable_col = None
    for index, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip().lower() == "suma variable":
            suma_variable_col = index
            break

    if suma_variable_col is not None:
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=suma_variable_col,
            max_col=suma_variable_col,
        ):
            row[0].number_format = "0.000%"

    aplicar_formato_encabezado(ws)
    autoajustar_columnas(ws)
    agregar_filtros(ws)
    wb.save(output_path)

    return output_path


def generar_consolidado_desde_response_json(
    response_json_path: str,
    output_dir: Optional[str] = None,
    filename: Optional[str] = None,
) -> Optional[str]:
    with open(response_json_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    consultas = payload.get("consultas") if isinstance(payload, dict) else None
    rows = normalize_srt_consulta_rows(consultas)
    target_dir = output_dir or os.path.dirname(response_json_path)
    return save_consolidated_excel(rows, target_dir, filename=filename)


def ejecutar_consulta_srt(
    base_url: str,
    api_key: str,
    email: str,
    cuit_login: str,
    clave: str,
    cuits_consulta: Any,
    proxy_request: Optional[bool],
    output_dir: str,
    raw_response_path: Optional[str] = None,
    log_fn: Optional[Callable[[str], None]] = None,
) -> Dict[str, Any]:
    response = consultar_srt_alicuotas(
        base_url=base_url,
        api_key=api_key,
        email=email,
        cuit_login=cuit_login,
        clave=clave,
        cuits_consulta=cuits_consulta,
        proxy_request=proxy_request,
    )

    data = response.get("data", {})
    consultas = data.get("consultas") if isinstance(data, dict) else None

    if raw_response_path:
        save_raw_response_json(data, raw_response_path)
        _log(f"Respuesta JSON guardada en {raw_response_path}", log_fn)

    json_paths = save_consultas_json_by_cuit(consultas, output_dir)
    if json_paths:
        _log(f"JSON por contribuyente guardados: {len(json_paths)}", log_fn)

    rows = normalize_srt_consulta_rows(consultas)
    excel_path = save_consolidated_excel(rows, output_dir)
    if excel_path:
        _log(f"Excel consolidado guardado en {excel_path}", log_fn)

    return {
        "response": response,
        "consultas": consultas,
        "rows": rows,
        "json_paths": json_paths,
        "excel_path": excel_path,
    }
