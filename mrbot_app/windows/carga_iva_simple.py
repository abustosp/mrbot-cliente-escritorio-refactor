import concurrent.futures
import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
import tkinter as tk
from tkinter import messagebox, ttk

from mrbot_app.carga_iva_simple import (
    CAMPOS_ARCHIVOS,
    carga_iva_simple,
    mapear_archivos_por_nombre,
    validar_archivos,
)
from mrbot_app.config import get_max_workers
from mrbot_app.helpers import df_preview, parse_bool_cell
from mrbot_app.windows.base import BaseWindow
from mrbot_app.windows.mixins import ExcelHandlerMixin, DownloadHandlerMixin

REPORT_BASE_DIR = os.path.join("descargas", "iva-simple-carga")

CAMPOS_BOOLEAN = [
    "operaciones_ng_o_e",
    "prorrateo_global",
    "prorrateo_asignacion_directa",
    "prorrateo_ambos",
    "importacion_definitiva_bienes",
    "importacion_servicios",
    "regimen_turiva",
    "bienes_usados",
]
CAMPOS_BOOLEAN_CON_NINGUNA = CAMPOS_BOOLEAN + ["ninguna_anteriores"]


def _parse_bool_celda(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return False
    return text in ("si", "sí", "1", "true", "yes", "y")


def _resolver_ninguna_anteriores(row: pd.Series) -> bool:
    for col in CAMPOS_BOOLEAN:
        if _parse_bool_celda(row.get(col)):
            return False
    val = row.get("ninguna_anteriores")
    if val is None:
        return True
    return _parse_bool_celda(val)


def _extraer_paths_fila(row: pd.Series) -> Dict[str, str]:
    archivos: Dict[str, str] = {}
    for col in CAMPOS_ARCHIVOS:
        filename = str(row.get(col, "")).strip()
        if not filename:
            continue
        dir_col = f"dir_{col}"
        directory = str(row.get(dir_col, "")).strip()
        if directory:
            full_path = os.path.join(directory, filename)
        else:
            full_path = filename
        archivos[col] = full_path
    return archivos


class CargaIvaSimpleWindow(BaseWindow, ExcelHandlerMixin, DownloadHandlerMixin):

    def __init__(self, master=None, config_provider=None, example_paths=None):
        super().__init__(master, title="Carga IVA Simple", config_provider=config_provider)
        ExcelHandlerMixin.__init__(self)
        DownloadHandlerMixin.__init__(self)
        try:
            self.iconbitmap(os.path.join("bin", "ABP-blanco-en-fondo-negro.ico"))
        except Exception:
            pass
        self.example_paths = example_paths or {}

        container = ttk.Frame(self, padding=10)
        container.pack(fill="both", expand=True)
        self.add_section_label(container, "Carga IVA Simple")
        self.add_info_label(
            container,
            "Carga de archivos TXT (LIV/LIC) y CSV de apertura (CF/DF) "
            "para la presentacion de IVA Simple en ARCA.\n"
            "Los archivos se envian como multipart. Se valida su existencia "
            "antes de ejecutar.",
        )

        inputs = ttk.Frame(container)
        inputs.pack(fill="x", pady=4)
        ttk.Label(inputs, text="CUIT Representante").grid(row=0, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(inputs, text="Clave Representante").grid(row=1, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(inputs, text="CUIT Representado").grid(row=2, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(inputs, text="Denominacion").grid(row=3, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(inputs, text="Periodo (AAAAMM)").grid(row=4, column=0, sticky="w", padx=4, pady=2)

        self.cuit_rep_var = tk.StringVar()
        self.clave_var = tk.StringVar()
        self.cuit_repr_var = tk.StringVar()
        self.denominacion_var = tk.StringVar()
        self.periodo_var = tk.StringVar()

        ttk.Entry(inputs, textvariable=self.cuit_rep_var, width=25).grid(row=0, column=1, padx=4, pady=2, sticky="ew")
        ttk.Entry(inputs, textvariable=self.clave_var, width=25, show="*").grid(row=1, column=1, padx=4, pady=2, sticky="ew")
        ttk.Entry(inputs, textvariable=self.cuit_repr_var, width=25).grid(row=2, column=1, padx=4, pady=2, sticky="ew")
        ttk.Entry(inputs, textvariable=self.denominacion_var, width=25).grid(row=3, column=1, padx=4, pady=2, sticky="ew")
        ttk.Entry(inputs, textvariable=self.periodo_var, width=15).grid(row=4, column=1, padx=4, pady=2, sticky="w")
        inputs.columnconfigure(1, weight=1)

        iva_opts = ttk.LabelFrame(container, text="Opciones IVA", padding=4)
        iva_opts.pack(fill="x", pady=4)
        self._bool_vars: Dict[str, tk.BooleanVar] = {}
        for col in CAMPOS_BOOLEAN:
            self._bool_vars[col] = tk.BooleanVar(value=False)
        self._ninguna_var = tk.BooleanVar(value=True)

        labels = {
            "operaciones_ng_o_e": "Op. No Grav. o Exentas",
            "prorrateo_global": "Prorrateo Global",
            "prorrateo_asignacion_directa": "Prorrateo Asig. Directa",
            "prorrateo_ambos": "Prorrateo Ambos",
            "importacion_definitiva_bienes": "Importac. Definitiva Bienes",
            "importacion_servicios": "Importac. Servicios",
            "regimen_turiva": "Regimen Turismo (TurIVA)",
            "bienes_usados": "Bienes Usados",
        }
        for i, col in enumerate(CAMPOS_BOOLEAN):
            row_i = i // 2
            col_i = i % 2
            ttk.Checkbutton(iva_opts, text=labels[col], variable=self._bool_vars[col]).grid(
                row=row_i, column=col_i, padx=4, pady=1, sticky="w"
            )
        ttk.Checkbutton(
            iva_opts, text="Ninguna de las anteriores", variable=self._ninguna_var
        ).grid(row=len(CAMPOS_BOOLEAN) // 2 + 1, column=0, columnspan=2, padx=4, pady=1, sticky="w")

        btns = ttk.Frame(container)
        btns.pack(fill="x", pady=8)
        ttk.Button(btns, text="Carga Individual", command=self.carga_individual).grid(
            row=0, column=0, padx=4, pady=2, sticky="ew"
        )
        ttk.Button(btns, text="Seleccionar Excel", command=self.cargar_excel).grid(
            row=0, column=1, padx=4, pady=2, sticky="ew"
        )
        ttk.Button(btns, text="Ver ejemplo", command=lambda: self.abrir_ejemplo_key("carga_iva_simple.xlsx")).grid(
            row=0, column=2, padx=4, pady=2, sticky="ew"
        )
        ttk.Button(btns, text="Previsualizar Excel", command=self.previsualizar_excel).grid(
            row=1, column=0, padx=4, pady=2, sticky="ew"
        )
        ttk.Button(btns, text="Validar archivos", command=self.validar_archivos_excel).grid(
            row=1, column=1, padx=4, pady=2, sticky="ew"
        )
        ttk.Button(btns, text="Procesar Excel", command=self.procesar_excel).grid(
            row=1, column=2, padx=4, pady=2, sticky="ew"
        )
        btns.columnconfigure((0, 1, 2), weight=1)

        self.preview = self.add_preview(container, height=8, show=False)
        self.set_preview(
            self.preview, "Selecciona un Excel y presiona 'Previsualizar Excel' para ver los datos."
        )

        self.progress_frame = self.add_progress_bar(container, label="Progreso")
        self.log_text = self.add_collapsible_log(
            container, title="Logs de ejecucion", height=16, service="carga_iva_simple"
        )

    def clear_logs(self) -> None:
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state="disabled")

    def append_log(self, text: str) -> None:
        if text:
            self.log_message(text)

    def _build_payload_from_ui(self) -> Dict[str, Any]:
        periodo = self.periodo_var.get().strip()
        periodo_digits = "".join(ch for ch in periodo if ch.isdigit())
        if len(periodo_digits) >= 6:
            periodo = periodo_digits[:6]
        payload: Dict[str, Any] = {
            "cuit_representante": self.cuit_rep_var.get().strip(),
            "clave_representante": self.clave_var.get(),
            "cuit_representado": self.cuit_repr_var.get().strip(),
            "denominacion": self.denominacion_var.get().strip(),
            "periodo": periodo,
        }
        for col in CAMPOS_BOOLEAN:
            payload[col] = bool(self._bool_vars[col].get())
        payload["ninguna_anteriores"] = bool(self._ninguna_var.get())
        return payload

    def carga_individual(self) -> None:
        payload = self._build_payload_from_ui()
        if not payload["cuit_representante"] or not payload["clave_representante"]:
            messagebox.showerror("Error", "Faltan datos obligatorios (CUIT y Clave).")
            return

        self.clear_logs()
        label = payload["cuit_representado"] or payload["cuit_representante"]
        self.log_start("Carga IVA Simple", {"modo": "individual", "cuit": label})

        self.run_in_thread(
            self.run_with_log_block,
            label,
            self._worker_individual,
            payload,
        )

    def _worker_individual(self, payload: Dict[str, Any]) -> None:
        label = payload.get("denominacion", "") or payload["cuit_representado"]
        self.log_separator(f"{label} ({payload['cuit_representado']})")

        response = carga_iva_simple(
            cuit_representante=payload["cuit_representante"],
            clave_representante=payload["clave_representante"],
            cuit_representado=payload["cuit_representado"],
            denominacion=payload["denominacion"],
            periodo=payload["periodo"],
            operaciones_ng_o_e=payload.get("operaciones_ng_o_e", False),
            prorrateo_global=payload.get("prorrateo_global", False),
            prorrateo_asignacion_directa=payload.get("prorrateo_asignacion_directa", False),
            prorrateo_ambos=payload.get("prorrateo_ambos", False),
            importacion_definitiva_bienes=payload.get("importacion_definitiva_bienes", False),
            importacion_servicios=payload.get("importacion_servicios", False),
            regimen_turiva=payload.get("regimen_turiva", False),
            bienes_usados=payload.get("bienes_usados", False),
            ninguna_anteriores=payload.get("ninguna_anteriores", True),
            log_fn=self.log_message,
        )

        http_status = response.get("http_status")
        success = response.get("success", http_status == 200)
        error = response.get("error") or ""
        detail = response.get("detail", "")
        detail_msg = ""
        if isinstance(detail, dict):
            detail_msg = "; ".join(str(m) for m in (detail.get("message") or []))
        elif isinstance(detail, str):
            detail_msg = detail
        self.log_info(f"HTTP {http_status} | success={success} | {error} {detail_msg}")

        self._guardar_log_individual(
            payload.get("cuit_representado", ""),
            payload.get("denominacion", ""),
            payload.get("periodo", ""),
            response,
        )

        self.log_info("Carga individual finalizada.")

    def validar_archivos_excel(self) -> None:
        if self.excel_df is None or self.excel_df.empty:
            messagebox.showerror("Error", "Carga un Excel primero.")
            return

        df_to_check = self._filter_procesar(self.excel_df)
        if df_to_check is None or df_to_check.empty:
            messagebox.showwarning("Sin filas", "No hay filas con procesar=SI")
            return

        self._run_file_validation(df_to_check)

    def _run_file_validation(self, df: pd.DataFrame) -> bool:
        all_missing: List[str] = []
        for idx, (_, row) in enumerate(df.iterrows(), start=2):
            archivos = _extraer_paths_fila(row)
            paths = list(archivos.values())
            missing = validar_archivos(paths)
            if missing:
                label = str(row.get("denominacion", "")) or str(row.get("cuit_representado", f"fila {idx}"))
                for p in missing:
                    all_missing.append(f"  {label}: {p}")

        if all_missing:
            msg = "Se encontraron archivos faltantes:\n\n" + "\n".join(all_missing)
            messagebox.showerror("Archivos faltantes", msg)
            return False
        else:
            messagebox.showinfo("Validacion exitosa", "Todos los archivos existen.")
            return True

    def procesar_excel(self) -> None:
        if self.excel_df is None or self.excel_df.empty:
            messagebox.showerror("Error", "Carga un Excel primero.")
            return

        df_to_process = self._filter_procesar(self.excel_df)
        if df_to_process is None or df_to_process.empty:
            messagebox.showwarning("Sin filas", "No hay filas con procesar=SI")
            return

        if not self._run_file_validation(df_to_process):
            return

        df_copy = df_to_process.copy()

        self.clear_logs()
        self.log_start("Carga IVA Simple", {"modo": "masivo", "filas": len(df_copy)})

        self.run_in_thread(self._worker_excel, df_copy)

    def _worker_excel(self, df: pd.DataFrame) -> None:
        rows: List[Dict[str, Any]] = []
        total = len(df)
        self.set_progress(0, total)
        max_workers = get_max_workers()

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(
                    self.run_with_log_block,
                    str(row.get("cuit_representado", "") or row.get("cuit_representante", "") or "sin_cuit").strip(),
                    self._process_row,
                    row,
                ): idx
                for idx, (_, row) in enumerate(df.iterrows(), start=1)
            }

            completed = 0
            for future in concurrent.futures.as_completed(futures):
                idx = futures[future]
                completed += 1
                if self._abort_event.is_set():
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

                try:
                    result = future.result()
                    if result:
                        rows.append(result)
                except Exception as exc:
                    self.log_error(f"Error en fila {idx}: {exc}")

                self.set_progress(completed, total)

        out_df = pd.DataFrame(rows)
        self.set_preview(self.preview, df_preview(out_df, rows=min(20, len(out_df))))
        success_count = sum(1 for r in rows if r.get("success"))
        error_count = sum(1 for r in rows if not r.get("success"))
        self.set_execution_summary({
            "source": "Carga IVA Simple",
            "total": max(total, 0),
            "success": success_count,
            "warning": max(0, total - success_count - error_count),
            "error": error_count,
        })

        self._generar_reporte_excel(rows)

        self.log_info("Procesamiento masivo finalizado.")

    def _process_row(self, row: pd.Series) -> Optional[Dict[str, Any]]:
        if self._abort_event.is_set():
            return None

        cuit_representante = str(row.get("cuit_representante", "")).strip()
        clave = str(row.get("clave_representante", "") or row.get("clave", "")).strip()
        cuit_representado = str(row.get("cuit_representado", "")).strip()
        denominacion = str(row.get("denominacion", "")).strip()
        periodo = str(row.get("periodo", "")).strip()
        periodo_digits = "".join(ch for ch in periodo if ch.isdigit())
        if len(periodo_digits) >= 6:
            periodo = periodo_digits[:6]

        ops: Dict[str, bool] = {}
        for col in CAMPOS_BOOLEAN:
            ops[col] = _parse_bool_celda(row.get(col))
        ops["ninguna_anteriores"] = _resolver_ninguna_anteriores(row)

        archivos = _extraer_paths_fila(row)

        self.log_separator(f"{denominacion} ({cuit_representado})")
        self.log_info(f"Archivos a cargar: {len(archivos)} ({', '.join(archivos.keys())})")
        activas = [k for k, v in ops.items() if v]
        self.log_info(f"Opciones activas: {', '.join(activas) if activas else 'ninguna'}")

        response = carga_iva_simple(
            cuit_representante=cuit_representante,
            clave_representante=clave,
            cuit_representado=cuit_representado,
            denominacion=denominacion,
            periodo=periodo,
            archivos=archivos,
            operaciones_ng_o_e=ops["operaciones_ng_o_e"],
            prorrateo_global=ops["prorrateo_global"],
            prorrateo_asignacion_directa=ops["prorrateo_asignacion_directa"],
            prorrateo_ambos=ops["prorrateo_ambos"],
            importacion_definitiva_bienes=ops["importacion_definitiva_bienes"],
            importacion_servicios=ops["importacion_servicios"],
            regimen_turiva=ops["regimen_turiva"],
            bienes_usados=ops["bienes_usados"],
            ninguna_anteriores=ops["ninguna_anteriores"],
            log_fn=self.log_message,
        )

        http_status = response.get("http_status")
        success = response.get("success", http_status == 200)
        error = str(response.get("error") or response.get("detail", "") or "")
        self.log_info(f"HTTP {http_status} | success={success} | {error}")

        self._guardar_log_individual(cuit_representado, denominacion, periodo, response)

        return {
            "cuit_representado": cuit_representado,
            "denominacion": denominacion,
            "periodo": periodo,
            "success": success,
            "http_status": http_status,
            "message": error,
            "archivos_cargados": len(archivos),
        }

    def _guardar_log_individual(self, cuit_representado: str, denominacion: str, periodo: str, response: Dict[str, Any]) -> None:
        try:
            cuit_dir = os.path.join(REPORT_BASE_DIR, cuit_representado)
            os.makedirs(cuit_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(cuit_dir, f"{timestamp}_response.json")
            log_data = {
                "cuit_representado": cuit_representado,
                "denominacion": denominacion,
                "periodo": periodo,
                "timestamp": datetime.now().isoformat(),
                "response": response,
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(log_data, f, indent=2, ensure_ascii=False, default=str)
            self.log_info(f"Log individual guardado: {path}")
        except Exception as exc:
            self.log_error(f"No se pudo guardar log individual: {exc}")

    def _generar_reporte_excel(self, rows: List[Dict[str, Any]]) -> None:
        try:
            os.makedirs(REPORT_BASE_DIR, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(REPORT_BASE_DIR, f"reporte-carga-iva-simple-{timestamp}.xlsx")

            filas = []
            for r in rows:
                filas.append({
                    "CUIT Representado": r.get("cuit_representado", ""),
                    "Denominacion": r.get("denominacion", ""),
                    "Periodo": r.get("periodo", ""),
                    "Success": "SI" if r.get("success") else "NO",
                    "HTTP Status": r.get("http_status") if r.get("http_status") is not None else "N/A",
                    "Archivos Cargados": r.get("archivos_cargados", 0),
                    "Error": str(r.get("message", ""))[:500],
                })
            df = pd.DataFrame(filas)

            total = len(rows)
            success_count = sum(1 for r in rows if r.get("success"))
            error_count = sum(1 for r in rows if not r.get("success"))
            df_resumen = pd.DataFrame({
                "Metrica": ["Total", "Exitosos", "Con error"],
                "Valor": [total, success_count, error_count],
            })

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Detalle", index=False)
                df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = load_workbook(path)
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
            wb.save(path)

            self.log_info(f"Reporte Excel guardado: {path}")
        except Exception as exc:
            self.log_error(f"No se pudo generar reporte Excel: {exc}")
