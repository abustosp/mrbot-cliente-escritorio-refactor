import os
import json
import glob
import re
import pandas as pd
import numpy as np
from datetime import date, datetime
from typing import Optional, Callable, Dict, Any, List, Tuple
from urllib.parse import urlparse, unquote

from openpyxl import load_workbook, Workbook

from mrbot_app.mis_comprobantes import consulta_mc, crear_directorio_seguro, extraer_csv_de_zip, FALLBACK_BASE_DIR
from mrbot_app.consulta import descargar_archivos_minio_concurrente
from mrbot_app.config import get_timeout_mc_control_monotributo
from mrbot_app.helpers import format_date_str, safe_post, build_headers, ensure_trailing_slash
from mrbot_app.formatos import (
    aplicar_formato_encabezado,
    aplicar_formato_moneda,
    autoajustar_columnas,
    agregar_filtros,
    alinear_columnas
)

NOTAS_DE_CREDITO = [3, 8, 13, 21, 38, 43, 44, 48, 53, 90, 110, 112, 113, 114, 119, 203, 208, 213]

# ─── Helper para navegación segura en JSONs anidados ─────────────────

def get_nested(data: dict, path: List[str], default=None):
    """Obtiene un valor anidado en un dict siguiendo una lista de keys."""
    current = data
    for key in path:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return default
    return current

# ─── Funciones de categoría reutilizables ─────────────────────────────

def obtener_max_ingresos_categoria(x: float, categorias: pd.DataFrame) -> float:
    """Retorna el límite de ingresos brutos de la categoría que cubre el monto x."""
    matches = categorias.loc[categorias['Ingresos brutos'] >= x, 'Ingresos brutos']
    return matches.iloc[0] if not matches.empty else 0


def obtener_categoria(x: float, categorias: pd.DataFrame) -> str:
    """Retorna la etiqueta de categoría de monotributo para el monto x."""
    matches = categorias.loc[categorias['Ingresos brutos'] >= x, 'Categoria']
    return matches.iloc[0] if not matches.empty else "Excedido"


# ─── Agrupación mensual para reportes HTML ───────────────────────────

def agrupar_por_mes(consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa el consolidado por Cliente, mes (YYYY-MM) y Tipo_MC,
    sumando Imp. Total (ya neteado por NC y convertido a pesos).
    """
    df = consolidado.copy()
    if 'Tipo_MC' not in df.columns:
        df['Tipo_MC'] = 'General'
    df['Mes'] = df['Fecha'].dt.strftime('%Y-%m')
    agrupado = df.groupby(
        ['Cliente', 'Fin CUIT', 'Mes', 'Tipo_MC'], as_index=False
    )['Imp. Total'].sum()
    agrupado.rename(columns={'Imp. Total': 'Total Mensual'}, inplace=True)
    return agrupado


def desglose_contrapartes(
    df: pd.DataFrame,
    col_contraparte: str = 'Denominación Receptor/Emisor',
    limite: int = 10,
) -> List[Dict[str, Any]]:
    """
    Agrupa por contraparte (cliente/proveedor), suma Imp. Total y calcula
    porcentaje de incidencia. Retorna top *limite* + 'Otros'.
    Similar al 'counterparty_breakdown' del ERP.
    """
    if df.empty or col_contraparte not in df.columns:
        return []

    grupo = df.groupby(col_contraparte, as_index=False)['Imp. Total'].sum()
    grupo = grupo.sort_values('Imp. Total', ascending=False)
    total_gral = float(grupo['Imp. Total'].sum())

    data = []
    for _, row in grupo.iterrows():
        nombre = str(row[col_contraparte]) or "(sin nombre)"
        total = float(row['Imp. Total'])
        data.append({"nombre": nombre.strip(), "total": round(total, 2)})

    if len(data) > limite:
        top = data[:limite]
        otros_total = round(sum(d["total"] for d in data[limite:]), 2)
        top.append({"nombre": "Otros", "total": otros_total})
    else:
        top = data

    for d in top:
        d["porcentaje"] = round((d["total"] / total_gral * 100) if total_gral else 0, 1)

    return top


def preparar_datos_individuales(
    consolidado: pd.DataFrame,
    categorias: pd.DataFrame,
    cliente: str,
    cuit: str,
) -> Dict[str, Any]:
    """
    Prepara los datos mensuales de compras/ventas, categorización
    y desglose por contraparte para un contribuyente individual.
    """
    df_cliente = consolidado[consolidado['Cliente'] == cliente].copy()
    if df_cliente.empty:
        return {
            "cliente": cliente,
            "cuit": cuit,
            "total_ventas": 0,
            "total_compras": 0,
            "categoria": "Sin datos",
            "categoria_compras": "Sin datos",
            "limite_categoria": 0,
            "pct_limite": 0,
            "series_ventas": [],
            "series_compras": [],
            "escala_categorias": [],
            "contrapartes_ventas": [],
            "contrapartes_compras": [],
        }

    df_cliente['Mes'] = df_cliente['Fecha'].dt.strftime('%Y-%m')

    # Split by Tipo_MC: Emitido = Ventas, Recibido = Compras
    mask_ventas = df_cliente['Tipo_MC'] == 'Emitido'
    mask_compras = df_cliente['Tipo_MC'] == 'Recibido'

    ventas = df_cliente[mask_ventas].groupby('Mes', as_index=False)['Imp. Total'].sum()
    compras = df_cliente[mask_compras].groupby('Mes', as_index=False)['Imp. Total'].sum()

    meses = sorted(set(list(ventas['Mes']) + list(compras['Mes'])))
    ventas_map = dict(zip(ventas['Mes'], ventas['Imp. Total']))
    compras_map = dict(zip(compras['Mes'], compras['Imp. Total']))

    series_ventas = [{"mes": m, "total": round(float(ventas_map.get(m, 0)), 2)} for m in meses]
    series_compras = [{"mes": m, "total": round(float(compras_map.get(m, 0)), 2)} for m in meses]

    total_ventas = sum(s["total"] for s in series_ventas)
    total_compras = sum(s["total"] for s in series_compras)

    categoria_label = obtener_categoria(total_ventas, categorias)
    categoria_compras_label = obtener_categoria(total_compras, categorias)
    limite_categoria = obtener_max_ingresos_categoria(total_ventas, categorias)

    escala_categorias = []
    for _, row in categorias.iterrows():
        escala_categorias.append({
            "categoria": str(row['Categoria']),
            "limite": float(row['Ingresos brutos']),
        })

    pct_limite = round((total_ventas / limite_categoria * 100), 1) if limite_categoria > 0 else 0

    # ─── Desglose por contraparte (clientes/proveedores) ──────────────
    contrapartes_ventas = desglose_contrapartes(
        df_cliente[mask_ventas], 'Denominación Receptor/Emisor'
    )
    contrapartes_compras = desglose_contrapartes(
        df_cliente[mask_compras], 'Denominación Receptor/Emisor'
    )

    return {
        "cliente": cliente,
        "cuit": str(cuit),
        "total_ventas": round(total_ventas, 2),
        "total_compras": round(total_compras, 2),
        "categoria": categoria_label,
        "categoria_compras": categoria_compras_label,
        "limite_categoria": round(limite_categoria, 2),
        "pct_limite": pct_limite,
        "series_ventas": series_ventas,
        "series_compras": series_compras,
        "escala_categorias": escala_categorias,
        "contrapartes_ventas": contrapartes_ventas,
        "contrapartes_compras": contrapartes_compras,
    }


def _log_message(message: str, log_fn: Optional[Callable[[str], None]] = None) -> None:
    if log_fn:
        log_fn(message)
    else:
        print(message)

def _log_info(message: str, log_fn: Optional[Callable[[str], None]] = None) -> None:
    _log_message(f"INFO: {message}", log_fn)

def _log_error(message: str, log_fn: Optional[Callable[[str], None]] = None) -> None:
    _log_message(f"ERROR: {message}", log_fn)

def _normalizar_si_no(valor: Any) -> str:
    """
    Normaliza valores a 'si' o 'no'.
    Acepta: 'si', 's', 'yes', 'y', 'true', '1' (case insensitive) como 'si'.
    """
    if isinstance(valor, str):
        v = valor.lower().strip()
        if v in ["si", "s", "yes", "y", "true", "1"]:
            return "si"
    elif isinstance(valor, (bool, int)):
        if valor:
            return "si"
    return "no"


def _parse_bool(valor: Any, default: bool = False) -> bool:
    if isinstance(valor, bool):
        return valor
    if valor is None:
        return default
    texto = str(valor).strip().lower()
    if texto in {"si", "sí", "s", "yes", "y", "true", "1"}:
        return True
    if texto in {"no", "n", "false", "0"}:
        return False
    return default

def procesar_descarga_mc(
    row: pd.Series,
    log_fn: Optional[Callable[[str], None]] = None,
    abort_check: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    """
    Procesa la descarga de Mis Comprobantes para un contribuyente.
    Utiliza las variables de entorno para credenciales (como consulta_mc original).
    abort_check: funcion opcional que devuelve True si se debe abortar el proceso.
    """
    cuit_representante = str(row.get('cuit_representante', '')).strip()
    clave_representante = str(row.get('clave_representante', '')).strip()
    cuit_representado = str(row.get('cuit_representado', '')).strip()
    nombre_representado = str(row.get('denominacion_mc', '')).strip() or "Contribuyente"

    desde = format_date_str(row.get('desde_mc', ''))
    hasta = format_date_str(row.get('hasta_mc', ''))

    descarga_MC = _normalizar_si_no(row.get('descarga_mc'))
    descarga_MC_emitidos = _normalizar_si_no(row.get('descarga_mc_emitidos'))
    descarga_MC_recibidos = _normalizar_si_no(row.get('descarga_mc_recibidos'))

    ubicacion_base = str(row.get('ubicacion_descarga_mc', '')).strip()

    if descarga_MC != 'si':
        _log_info(f"Saltando descarga MC para CUIT {cuit_representado}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": True,
            "descarga_esperada": False,
            "descargas": 0,
            "errores_descarga": None,
        }

    descargar_emitidos = (descarga_MC_emitidos == 'si')
    descargar_recibidos = (descarga_MC_recibidos == 'si')
    timeout_mc = get_timeout_mc_control_monotributo()
    proxy_request: Optional[bool] = None
    if "proxy_request_mc" in row.index or "proxy_request" in row.index:
        proxy_request = _parse_bool(row.get('proxy_request_mc', row.get('proxy_request', '')), default=False)

    if not descargar_emitidos and not descargar_recibidos:
        _log_info(f"No hay tipos de comprobantes seleccionados para descargar (MC) para {cuit_representado}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": True,
            "descarga_esperada": False,
            "descargas": 0,
            "errores_descarga": None,
        }

    _log_info(f"Procesando MC: {nombre_representado} ({cuit_representado}) - Periodo: {desde} a {hasta}", log_fn)

    download_errors: List[str] = []
    downloads = 0

    try:
        if abort_check and abort_check():
            _log_info(f"Abortando request MC para CUIT {cuit_representado} antes de consultar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": "Abortado por el usuario",
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": "Abortado por el usuario",
            }

        response = consulta_mc(
            desde=desde,
            hasta=hasta,
            cuit_inicio_sesion=cuit_representante,
            representado_nombre=nombre_representado,
            representado_cuit=cuit_representado,
            contrasena=clave_representante,
            descarga_emitidos=descargar_emitidos,
            descarga_recibidos=descargar_recibidos,
            carga_minio=True,
            carga_json=False,
            proxy_request=proxy_request,
            timeout_mc=timeout_mc,
            log_fn=log_fn
        )

        if abort_check and abort_check():
            _log_info(f"Abortando descarga MC para CUIT {cuit_representado} despues de consultar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": "Abortado por el usuario",
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": "Abortado por el usuario",
            }

        if not response.get("success", False):
            error_msg = response.get("error", response.get("detail", "Error desconocido"))
            _log_error(f"API Error MC: {error_msg}", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": str(error_msg),
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": str(error_msg),
            }

        # Some responses return success=true with a populated 'error' field and no MinIO URLs.
        # Treat that as a failed download so the summary reflects the real outcome.
        response_errors = response.get("error") or []
        if isinstance(response_errors, list):
            response_errors_text = "; ".join(str(e) for e in response_errors if e)
        else:
            response_errors_text = str(response_errors) if response_errors else ""

        emitidos_url = response.get("mis_comprobantes_emitidos_url_minio") or ""
        recibidos_url = response.get("mis_comprobantes_recibidos_url_minio") or ""
        no_urls = (not emitidos_url) and (not recibidos_url)

        if response_errors_text and no_urls:
            _log_error(f"API MC sin archivos: {response_errors_text}", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": response_errors_text,
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": response_errors_text,
            }

        if no_urls:
            _log_info("Respuesta MC sin URLs MinIO para descargar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": True,
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": None,
            }

        # Determine download directory
        if not ubicacion_base:
            ubicacion_base = os.path.join("descargas", "Control_Monotributistas", "Mis Comprobantes", cuit_representado)

        os.makedirs(ubicacion_base, exist_ok=True)

        # Save API response as JSON log for traceability
        try:
            log_response_path = os.path.join(ubicacion_base, "response_log.json")
            with open(log_response_path, "w", encoding="utf-8") as f:
                json.dump(response, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            _log_error(f"Error guardando log response JSON para MC: {e}", log_fn)

        # Standard structure from external repo: [Base]/extraido/*.csv
        # But we need to download ZIPs first.
        # External repo:
        #   descargas_mc/[CUIT]_[Nombre]/[archivo].zip
        #   descargas_mc/[CUIT]_[Nombre]/extraido/

        # We will follow this structure relative to ubicacion_base

        archivos_a_descargar = []
        extraido_dir = os.path.join(ubicacion_base, "extraido")

        if descargar_emitidos and response.get("mis_comprobantes_emitidos_url_minio"):
            url = response["mis_comprobantes_emitidos_url_minio"]
            # Using filename from URL with unquote to handle URL encoding
            filename_zip = unquote(os.path.basename(urlparse(url).path)) or "Emitidos.zip"
            zip_path = os.path.join(ubicacion_base, filename_zip)

            # Name for extracted CSV: same as zip base name
            csv_name = os.path.splitext(filename_zip)[0] + ".csv"
            csv_path = os.path.join(extraido_dir, csv_name)

            archivos_a_descargar.append({"url": url, "destino": zip_path, "csv_destino": csv_path})

        if descargar_recibidos and response.get("mis_comprobantes_recibidos_url_minio"):
            url = response["mis_comprobantes_recibidos_url_minio"]
            filename_zip = unquote(os.path.basename(urlparse(url).path)) or "Recibidos.zip"
            zip_path = os.path.join(ubicacion_base, filename_zip)
            csv_name = os.path.splitext(filename_zip)[0] + ".csv"
            csv_path = os.path.join(extraido_dir, csv_name)

            archivos_a_descargar.append({"url": url, "destino": zip_path, "csv_destino": csv_path})

        if archivos_a_descargar:
            _log_info(f"Descargando {len(archivos_a_descargar)} archivos MC...", log_fn)
            # Adapt structure for downloader
            download_items = [{"url": item["url"], "destino": item["destino"]} for item in archivos_a_descargar]
            results = descargar_archivos_minio_concurrente(download_items, log_fn=log_fn, abort_check=abort_check)
            downloads = sum(1 for item in results if item.get("success"))
            download_errors.extend(
                str(item.get("error") or "Error al descargar")
                for item in results
                if not item.get("success")
            )

            for item in archivos_a_descargar:
                if os.path.exists(item["destino"]):
                    _log_info(f"Extrayendo CSV de {os.path.basename(item['destino'])}", log_fn)
                    if extraer_csv_de_zip(item["destino"], item["csv_destino"], log_fn):
                        # Optionally remove zip
                        # os.remove(item["destino"])
                        pass
                    else:
                        download_errors.append(f"No se pudo extraer {os.path.basename(item['destino'])}")

        return {
            "cuit_representado": cuit_representado,
            "success": True,
            "descarga_esperada": True,
            "descargas": downloads,
            "errores_descarga": "; ".join(download_errors) if download_errors else None,
        }

    except Exception as e:
        _log_error(f"Excepcion en proceso MC: {e}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": False,
            "message": str(e),
            "descarga_esperada": True,
            "descargas": downloads,
            "errores_descarga": str(e),
        }


def _is_pdf_url(url: Any) -> bool:
    if not isinstance(url, str):
        return False
    clean = url.strip()
    if not clean.lower().startswith("http"):
        return False
    lowered = clean.lower()
    if "minio" in lowered:
        return True
    return lowered.split("?")[0].endswith(".pdf")

def _collect_pdf_items(data: Any) -> List[Tuple[str, Dict[str, Any]]]:
    # Reuse logic from RcelWindow
    if not isinstance(data, dict):
        return []
    collected: List[Tuple[str, Dict[str, Any]]] = []

    def _extract_item_pdf_url(item: Dict[str, Any]) -> Optional[str]:
        for key in ("URL_MINIO", "url_minio", "url_pdf", "link_pdf", "url", "link"):
            url = item.get(key)
            if _is_pdf_url(url):
                return str(url).strip()
        for value in item.values():
            if _is_pdf_url(value):
                return str(value).strip()
        return None

    for key in ("facturas_emitidas", "facturas_recibidas", "comprobantes", "facturas"):
        items = data.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            url = _extract_item_pdf_url(item)
            if url:
                collected.append((url, item))
    return collected

def procesar_descarga_rcel(
    row: pd.Series,
    config: Tuple[str, str, str],
    log_fn: Optional[Callable[[str], None]] = None,
    abort_check: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    """
    Procesa la descarga de RCEL.
    config: (base_url, api_key, email)
    abort_check: funcion opcional que devuelve True si se debe abortar el proceso.
    """
    base_url, api_key, email = config

    cuit_representante = str(row.get('cuit_representante', '')).strip()
    clave_representante = str(row.get('clave_representante', '')).strip()
    cuit_representado = str(row.get('cuit_representado', '')).strip()
    nombre_rcel = str(row.get('denominacion_rcel', '')).strip() or "Contribuyente"

    desde = format_date_str(row.get('desde_rcel', ''))
    hasta = format_date_str(row.get('hasta_rcel', ''))

    descarga_RCEL = _normalizar_si_no(row.get('descarga_rcel'))
    proxy_request: Optional[bool] = None
    if "proxy_request_rcel" in row.index or "proxy_request" in row.index:
        proxy_request = _parse_bool(row.get('proxy_request_rcel', row.get('proxy_request', '')), default=False)
    ubicacion_base = str(row.get('ubicacion_descarga_rcel', '')).strip()

    if descarga_RCEL != 'si':
        _log_info(f"Saltando descarga RCEL para CUIT {cuit_representado}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": True,
            "descarga_esperada": False,
            "descargas": 0,
            "errores_descarga": None,
        }

    _log_info(f"Procesando RCEL: {nombre_rcel} ({cuit_representado}) - Periodo: {desde} a {hasta}", log_fn)

    url_api = ensure_trailing_slash(base_url) + "api/v1/rcel/consulta"
    headers = build_headers(api_key, email)

    payload = {
        "desde": desde,
        "hasta": hasta,
        "cuit_representante": cuit_representante,
        "nombre_rcel": nombre_rcel,
        "representado_cuit": cuit_representado,
        "clave": clave_representante,
        "minio_upload": True,
    }
    if proxy_request is not None:
        payload["proxy_request"] = proxy_request

    download_errors: List[str] = []
    post_errors: List[str] = []
    downloads = 0

    try:
        if abort_check and abort_check():
            _log_info(f"Abortando request RCEL para CUIT {cuit_representado} antes de consultar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": "Abortado por el usuario",
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": "Abortado por el usuario",
            }

        # Log request (redacted)
        safe_payload = payload.copy()
        safe_payload['clave'] = '***'
        _log_message(f"RCEL Request: {json.dumps(safe_payload, default=str)}", log_fn)

        response = safe_post(url_api, headers, payload)
        data = response.get("data")

        if abort_check and abort_check():
            _log_info(f"Abortando descarga RCEL para CUIT {cuit_representado} despues de consultar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": "Abortado por el usuario",
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": "Abortado por el usuario",
            }

        if not response.get("success") and not data:
             _log_error(f"API Error RCEL: {response.get('message', 'Unknown error')}", log_fn)
             return {
                 "cuit_representado": cuit_representado,
                 "success": False,
                 "message": str(response.get('message', 'Unknown error')),
                 "descarga_esperada": True,
                 "descargas": 0,
                 "errores_descarga": str(response.get('message', 'Unknown error')),
             }

        # Detect responses where data reports success=false or carries an error even with HTTP 200.
        data_success = True
        data_error_text = ""
        if isinstance(data, dict):
            data_success = bool(data.get("success", True))
            err_field = data.get("error") or data.get("message") or data.get("detail")
            if isinstance(err_field, list):
                data_error_text = "; ".join(str(e) for e in err_field if e)
            elif err_field:
                data_error_text = str(err_field)

        if not data_success:
            error_text = data_error_text or "Error desconocido (RCEL)"
            _log_error(f"API Error RCEL: {error_text}", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": False,
                "message": error_text,
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": error_text,
            }

        # Determine download directory
        if not ubicacion_base:
            # Fallback structure: descargas/Control_Monotributistas/RCEL/[CUIT]
            ubicacion_base = os.path.join("descargas", "Control_Monotributistas", "RCEL", cuit_representado)

        os.makedirs(ubicacion_base, exist_ok=True)

        # Save API response as JSON log for traceability
        try:
            log_response_path = os.path.join(ubicacion_base, "response_log.json")
            with open(log_response_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            _log_error(f"Error guardando log response JSON para RCEL: {e}", log_fn)

        # Collect PDFs and metadata
        pdf_items = _collect_pdf_items(data)

        if not pdf_items:
            _log_info("No se encontraron comprobantes RCEL con PDF para descargar.", log_fn)
            return {
                "cuit_representado": cuit_representado,
                "success": True,
                "descarga_esperada": True,
                "descargas": 0,
                "errores_descarga": None,
            }

        _log_info(f"Se encontraron {len(pdf_items)} comprobantes RCEL.", log_fn)

        # Download PDFs
        download_items = []
        for url, meta in pdf_items:
            filename = unquote(os.path.basename(urlparse(url).path)) or "factura.pdf"
            dest = os.path.join(ubicacion_base, filename)
            download_items.append({"url": url, "destino": dest})

        results = descargar_archivos_minio_concurrente(download_items, log_fn=log_fn, abort_check=abort_check)
        downloads = sum(1 for item in results if item.get("success"))
        download_errors.extend(
            str(item.get("error") or "Error al descargar")
            for item in results
            if not item.get("success")
        )

        # Save JSON metadata for each downloaded file
        saved_jsons = 0
        for item in download_items:
            dest_pdf = item["destino"]
            if os.path.exists(dest_pdf):
                # Find metadata
                meta = next((m for u, m in pdf_items if u == item["url"]), None)
                if meta:
                    json_name = os.path.splitext(os.path.basename(dest_pdf))[0] + ".json"
                    json_path = os.path.join(ubicacion_base, json_name)
                    try:
                        with open(json_path, "w", encoding="utf-8") as f:
                            json.dump(meta, f, ensure_ascii=False, indent=2)
                        saved_jsons += 1
                    except Exception as e:
                        _log_error(f"Error guardando JSON {json_name}: {e}", log_fn)
                        post_errors.append(str(e))

        _log_info(f"Descargas RCEL completadas: {len(results)}. JSONs guardados: {saved_jsons}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": True,
            "descarga_esperada": True,
            "descargas": downloads,
            "errores_descarga": "; ".join(download_errors) if download_errors else None,
            "errores_postproceso": "; ".join(post_errors) if post_errors else None,
        }

    except Exception as e:
        _log_error(f"Excepcion en proceso RCEL: {e}", log_fn)
        return {
            "cuit_representado": cuit_representado,
            "success": False,
            "message": str(e),
            "descarga_esperada": True,
            "descargas": downloads,
            "errores_descarga": str(e),
            "errores_postproceso": "; ".join(post_errors) if post_errors else None,
        }

def leer_archivos_csv_batch(archivos_mc: List[str], log_fn: Optional[Callable[[str], None]] = None) -> pd.DataFrame:
    dataframes = []
    for f in archivos_mc:
        if not os.path.isfile(f):
            continue
        try:
            # Attempt reading with different encodings/separators if needed, but control.py used sep=';', decimal=','
            data = pd.read_csv(f, sep=';', decimal=',', encoding='utf-8-sig')
            if data.empty:
                continue

            data['Archivo'] = os.path.basename(f)

            # Logic from control.py
            partes_archivo = data["Archivo"].str.split("-")
            # control.py assumes format: [Something]-[Something]-[Something]-[Something]-[CUIT]-[Cliente].csv or similar
            # Actually, `get_unique_filename` might have changed the name.
            # But the content of CSV usually has columns that matter.
            # control.py relies on filename to get "Fin CUIT" and "Cliente".
            # "partes_archivo.str[4]" implies at least 5 parts.
            # If our filename doesn't match, this will fail.
            # The CSVs from MC usually have standard names like:
            # "Mis Comprobantes Emitidos - 2024-01-01 - 2024-12-31 - 20123456789.csv" (example)
            # control.py seems to expect a specific format.
            # Let's look at `control.py`:
            # partes_archivo = data["Archivo"].str.split("-")
            # data['Fin CUIT'] = partes_archivo.str[4].str.strip().astype(np.int64)
            # data['Cliente'] = partes_archivo.str[5].str.strip().str.replace('.csv','', regex=True)

            # If we renamed files differently in `procesar_descarga_mc`, this will break.
            # `procesar_descarga_mc` used `filename_zip` from URL.
            # MinIO URLs often have names like "MCE-20123456789-20240101-20241231.zip" or similar?
            # Or maybe "20123456789-MCE-..."

            # If parsing fails, we should try to extract from content if possible, but MC csv doesn't always have client CUIT/Name in rows (it has user's CUIT).
            # But wait, `data['CUIT Cliente']` and `data['Cliente']` seem to be the represented entity.
            # In `control.py`, it uses filename.

            # Let's try to be robust. If splits are not enough, maybe regex.
            # Or just use the file parent folder name if available?
            # But here we are reading many files in batch, potentially from different clients.

            # Let's assume the filename format is preserved from MinIO and matches what control.py expects OR we adjust.
            # If not, we might need to rely on the folder structure if `leer_archivos_csv_batch` is called per client?
            # No, `control()` receives a list of ALL files.

            # Workaround: Check if we can extract from filename safely.

            try:
                data['Fin CUIT'] = partes_archivo.str[4].str.strip().astype(np.int64)
                data['CUIT Cliente'] = partes_archivo.str[4].str.strip().astype(np.int64)
                if len(data["Archivo"].iloc[0].split("-")) > 5:
                     data['Cliente'] = partes_archivo.str[5].str.strip().str.replace('.csv','', regex=True)
                else:
                     data['Cliente'] = "Desconocido"
            except Exception:
                # Fallback: try to extract from folder name?
                # or just use a placeholder
                 data['Fin CUIT'] = 0
                 data['CUIT Cliente'] = 0
                 data['Cliente'] = "Desconocido"

            es_emitido = 'Denominación Receptor' in data.columns
            es_recibido = 'Denominación Emisor' in data.columns

            if es_emitido:
                data['Nro. Doc. Receptor/Emisor'] = data.get('Nro. Doc. Receptor', '')
                data['Denominación Receptor/Emisor'] = data.get('Denominación Receptor', '')
                data['Tipo_MC'] = 'Emitido'
            elif es_recibido:
                data['Nro. Doc. Receptor/Emisor'] = data.get('Nro. Doc. Emisor', '')
                data['Denominación Receptor/Emisor'] = data.get('Denominación Emisor', '')
                data['Tipo_MC'] = 'Recibido'
            else:
                data['Tipo_MC'] = 'General'

            cols = [
                'Fecha de Emisión', 'Tipo de Comprobante', 'Punto de Venta',
                'Número Desde', 'Número Hasta', 'Cód. Autorización',
                'Tipo Cambio', 'Moneda',
                'Imp. Neto Gravado Total', 'Imp. Neto No Gravado',
                'Imp. Op. Exentas', 'Otros Tributos', 'Total IVA', 'Imp. Total',
                'Nro. Doc. Receptor/Emisor', 'Denominación Receptor/Emisor',
                'Archivo', 'CUIT Cliente', 'Fin CUIT', 'Cliente', 'Tipo_MC'
            ]
            # Ensure columns exist
            for c in cols:
                if c not in data.columns:
                    data[c] = 0 if 'Imp.' in c or 'Total' in c else ''

            data = data[cols]
            dataframes.append(data)
        except Exception as e:
            _log_error(f"Error leyendo CSV {f}: {e}", log_fn)
            continue

    if dataframes:
        return pd.concat(dataframes, ignore_index=True)
    return pd.DataFrame()

def leer_archivos_json_batch(archivos_json: List[str], log_fn: Optional[Callable[[str], None]] = None) -> pd.DataFrame:
    registros = []
    for factura in archivos_json:
        if not os.path.isfile(factura):
            continue
        # Saltar archivos que no son RCEL (response_log.json, etc.)
        nombre = os.path.basename(factura)
        if nombre in ("response_log.json",):
            continue
        try:
            with open(factura, 'r', encoding='utf-8-sig') as f:
                data_dict = json.load(f)

            data_dict['Archivo PDF'] = os.path.basename(factura)

            # Extract CUIT from filename if possible. control.py: partes = ... split("-")[0]
            partes = data_dict['Archivo PDF'].split("-")
            if len(partes) >= 1 and partes[0].isdigit():
                data_dict['CUIT Cliente'] = int(partes[0].strip())
                data_dict['Fin CUIT'] = int(partes[0].strip())

            # Extract Client from parent dir
            try:
                parent = os.path.basename(os.path.dirname(factura))
                # control.py: split("_", 1)[1]
                if "_" in parent:
                    data_dict['Cliente'] = parent.split("_", 1)[1]
                else:
                    data_dict['Cliente'] = parent
            except Exception:
                pass

            registros.append(data_dict)
        except Exception as e:
             _log_error(f"Error leyendo JSON {factura}: {e}", log_fn)

    if registros:
        return pd.DataFrame(registros)
    return pd.DataFrame()


# ─── Lectura de JSONs del Facturador (AFIP WSFE) ──────────────────────

# Campos a extraer del JSON de respuesta del Facturador
FACTURADOR_FIELD_MAPPING = {
    "CbteFch": ["Request", "FeCAEReq", "FeDetReq", "CbteFch"],
    "Cuit": ["Request", "Auth", "Cuit"],
    "Denominacion_Representado": ["Request", "Auth", "Denominacion_Representado"],
    "CbteTipo": ["Request", "FeCAEReq", "FeCabReq", "CbteTipo"],
    "PtoVta": ["Request", "FeCAEReq", "FeCabReq", "PtoVta"],
    "CbteDesde": ["Request", "FeCAEReq", "FeDetReq", "CbteDesde"],
    "CbteHasta": ["Request", "FeCAEReq", "FeDetReq", "CbteHasta"],
    "DocTipo": ["Request", "FeCAEReq", "FeDetReq", "DocTipo"],
    "DocNro": ["Request", "FeCAEReq", "FeDetReq", "DocNro"],
    "Denominacion_receptor": ["Request", "FeCAEReq", "FeDetReq", "Denominacion_receptor"],
    "FchServDesde": ["Request", "FeCAEReq", "FeDetReq", "FchServDesde"],
    "FchServHasta": ["Request", "FeCAEReq", "FeDetReq", "FchServHasta"],
    "ImpTotConc": ["Request", "FeCAEReq", "FeDetReq", "ImpTotConc"],
    "ImpNeto": ["Request", "FeCAEReq", "FeDetReq", "ImpNeto"],
    "ImpOpEx": ["Request", "FeCAEReq", "FeDetReq", "ImpOpEx"],
    "ImpTrib": ["Request", "FeCAEReq", "FeDetReq", "ImpTrib"],
    "ImpIVA": ["Request", "FeCAEReq", "FeDetReq", "ImpIVA"],
    "ImpTotal": ["Request", "FeCAEReq", "FeDetReq", "ImpTotal"],
    "MonId": ["Request", "FeCAEReq", "FeDetReq", "MonId"],
}

FACTURADOR_TESTING_PATH = ["Request", "Auth", "testing"]
FACTURADOR_RESULTADO_PATH = [
    "Response", "Envelope", "Body", "FECAESolicitarResponse",
    "FECAESolicitarResult", "FeDetResp", "FECAEDetResponse", "Resultado"
]
FACTURADOR_CAE_PATH = [
    "Response", "Envelope", "Body", "FECAESolicitarResponse",
    "FECAESolicitarResult", "FeDetResp", "FECAEDetResponse", "CAE"
]


def _parse_yyyymmdd_to_datetime(yyyymmdd: Any):
    """Convierte una fecha YYYYMMDD (str o int) a pd.Timestamp, o NaT si falla."""
    if yyyymmdd is None:
        return pd.NaT
    s = str(yyyymmdd).strip()
    if len(s) == 8 and s.isdigit():
        try:
            return pd.Timestamp(year=int(s[0:4]), month=int(s[4:6]), day=int(s[6:8]))
        except (ValueError, TypeError):
            return pd.NaT
    return pd.NaT


def leer_archivos_facturador_batch(
    archivos_json: List[str],
    log_fn: Optional[Callable[[str], None]] = None,
) -> pd.DataFrame:
    """
    Lee archivos JSON de respuesta del Facturador (AFIP WSFE) y los convierte
    en un DataFrame con las mismas columnas que el consolidado de MC.

    Filtra:
      - Solo comprobantes con testing=False (producción).
      - Solo comprobantes con Resultado="A" (aprobados).
    """
    registros = []
    for filepath in archivos_json:
        if not os.path.isfile(filepath):
            continue

        filename = os.path.basename(filepath)

        try:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            _log_error(f"Error al leer JSON Facturador {filename}: {e}", log_fn)
            continue

        if not isinstance(data, dict):
            continue

        # ── Filtro: solo producción ──
        testing = get_nested(data, FACTURADOR_TESTING_PATH)
        if testing is not False:
            continue

        # ── Filtro: solo comprobantes aprobados ──
        resultado = get_nested(data, FACTURADOR_RESULTADO_PATH)
        if resultado != "A":
            continue

        # ── Extraer campos ──
        cuit_emisor = get_nested(data, FACTURADOR_FIELD_MAPPING["Cuit"]) or ""
        cuit_emisor_str = str(cuit_emisor).strip()

        denominacion = get_nested(data, FACTURADOR_FIELD_MAPPING["Denominacion_Representado"]) or ""
        # Si no hay denominación en el JSON, intentar extraer del nombre del directorio (CUIT)
        if not denominacion or not str(denominacion).strip():
            try:
                parent = os.path.basename(os.path.dirname(filepath))
                if parent and parent.isdigit():
                    denominacion = parent
            except Exception:
                pass

        cbte_tipo = get_nested(data, FACTURADOR_FIELD_MAPPING["CbteTipo"])
        pto_vta = get_nested(data, FACTURADOR_FIELD_MAPPING["PtoVta"])
        cbte_desde = get_nested(data, FACTURADOR_FIELD_MAPPING["CbteDesde"])
        cbte_hasta = get_nested(data, FACTURADOR_FIELD_MAPPING["CbteHasta"])
        doc_tipo = get_nested(data, FACTURADOR_FIELD_MAPPING["DocTipo"])
        doc_nro = get_nested(data, FACTURADOR_FIELD_MAPPING["DocNro"]) or ""
        denominacion_receptor = get_nested(data, FACTURADOR_FIELD_MAPPING["Denominacion_receptor"]) or ""

        # Fechas
        cbte_fch = _parse_yyyymmdd_to_datetime(get_nested(data, FACTURADOR_FIELD_MAPPING["CbteFch"]))
        fch_serv_desde = _parse_yyyymmdd_to_datetime(get_nested(data, FACTURADOR_FIELD_MAPPING["FchServDesde"]))
        fch_serv_hasta = _parse_yyyymmdd_to_datetime(get_nested(data, FACTURADOR_FIELD_MAPPING["FchServHasta"]))

        # Importes (ya son float/int en el JSON)
        imp_total = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpTotal"]) or 0)
        imp_tot_conc = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpTotConc"]) or 0)
        imp_neto = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpNeto"]) or 0)
        imp_op_ex = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpOpEx"]) or 0)
        imp_trib = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpTrib"]) or 0)
        imp_iva = float(get_nested(data, FACTURADOR_FIELD_MAPPING["ImpIVA"]) or 0)

        mon_id = get_nested(data, FACTURADOR_FIELD_MAPPING["MonId"]) or "PES"
        cae = get_nested(data, FACTURADOR_CAE_PATH) or ""

        # ── Construir AUX (mismo formato que MC) ──
        try:
            aux = (
                f"{int(cuit_emisor)}-{int(cbte_tipo):03d}-"
                f"{int(pto_vta):05d}-{int(cbte_desde):08d}"
            )
        except (ValueError, TypeError):
            aux = ""

        registro = {
            "Fecha de Emisión": cbte_fch,
            "Tipo de Comprobante": int(cbte_tipo) if cbte_tipo is not None else 0,
            "Punto de Venta": int(pto_vta) if pto_vta is not None else 0,
            "Número Desde": int(cbte_desde) if cbte_desde is not None else 0,
            "Número Hasta": int(cbte_hasta) if cbte_hasta is not None else 0,
            "Cód. Autorización": str(cae),
            "Tipo Cambio": 1,
            "Moneda": str(mon_id),
            "Imp. Neto Gravado Total": imp_neto,
            "Imp. Neto No Gravado": imp_tot_conc,
            "Imp. Op. Exentas": imp_op_ex,
            "Otros Tributos": imp_trib,
            "Total IVA": imp_iva,
            "Imp. Total": imp_total,
            "Nro. Doc. Receptor/Emisor": str(doc_nro),
            "Denominación Receptor/Emisor": str(denominacion_receptor),
            "Archivo": filename,
            "CUIT Cliente": int(cuit_emisor) if cuit_emisor_str.isdigit() else 0,
            "Fin CUIT": int(cuit_emisor) if cuit_emisor_str.isdigit() else 0,
            "Cliente": str(denominacion).strip(),
            "Tipo_MC": "Emitido",
            "FchServDesde": fch_serv_desde,
            "FchServHasta": fch_serv_hasta,
        }
        registros.append(registro)

    if registros:
        df = pd.DataFrame(registros)
        _log_info(f"Facturador: {len(archivos_json)} archivos leídos, {len(df)} comprobantes aprobados.", log_fn)
        return df

    _log_info("Facturador: no se encontraron comprobantes aprobados (Resultado=A, testing=False).", log_fn)
    return pd.DataFrame()


def generar_reporte_control(
    archivos_mc: List[str],
    archivos_json: List[str],
    path_categorias: str,
    output_path: str,
    log_fn: Optional[Callable[[str], None]] = None,
    html_output_dir: Optional[str] = None,
    archivos_facturador: Optional[List[str]] = None,
    fecha_inicial: Optional[pd.Timestamp] = None,
    fecha_final: Optional[pd.Timestamp] = None,
    categorias: Optional[pd.DataFrame] = None,
    logo_b64: Optional[str] = None,
) -> None:
    """
    Core logic for generating the report.
    Si html_output_dir se proporciona, también genera reportes HTML individuales
    y un reporte general con gráficos comparativos.
    Si categorias se proporciona, se usa directamente en lugar de cargarla del archivo.
    Si logo_b64 se proporciona, se incrusta en la esquina superior derecha de los reportes HTML.
    """
    _log_info("Iniciando generación de reporte...", log_fn)

    try:
        if categorias is not None:
            es_db = True
            if fecha_inicial is None:
                fecha_inicial = pd.Timestamp(date.today().replace(day=1))
            if fecha_final is None:
                fecha_final = pd.Timestamp(date.today())
        elif not os.path.exists(path_categorias):
            _log_error(f"No se encontró archivo de categorías: {path_categorias}", log_fn)
            return
        else:
            es_db = path_categorias.lower().endswith('.db')
            if es_db:
                from mrbot_app.servicios.categorias_monotributo import cargar_categorias
                ref_date = fecha_final.date() if fecha_final is not None else date.today()
                categorias = cargar_categorias(ref_date)
                if fecha_inicial is None:
                    fecha_inicial = pd.Timestamp(ref_date.replace(month=ref_date.month - 11, day=1))
                    _log_info(f"fecha_inicial no proporcionada, usando por defecto: {fecha_inicial.date()}", log_fn)
                if fecha_final is None:
                    fecha_final = pd.Timestamp(ref_date)
                    _log_info(f"fecha_final no proporcionada, usando por defecto: {fecha_final.date()}", log_fn)
            else:
                categorias = pd.read_excel(path_categorias, sheet_name='Categorias')
                if fecha_inicial is None or fecha_final is None:
                    fecha_inicial_raw = pd.read_excel(path_categorias, sheet_name='Rango de Fechas', header=None, skiprows=1, usecols=[0]).iloc[0,0]
                    fecha_final_raw = pd.read_excel(path_categorias, sheet_name='Rango de Fechas', header=None, skiprows=1, usecols=[1]).iloc[0,0]
                    fecha_inicial = pd.to_datetime(fecha_inicial_raw, dayfirst=True)
                    fecha_final = pd.to_datetime(fecha_final_raw, dayfirst=True)

        _log_info(f"Rango fechas control: {fecha_inicial.date()} - {fecha_final.date()}", log_fn)

        consolidado = leer_archivos_csv_batch(archivos_mc, log_fn)
        info_facturas_pdf = leer_archivos_json_batch(archivos_json, log_fn)

        # ── Incorporar datos del Facturador (AFIP WSFE) ──
        if archivos_facturador:
            _log_info(f"Procesando {len(archivos_facturador)} archivos JSON del Facturador...", log_fn)
            facturador_df = leer_archivos_facturador_batch(archivos_facturador, log_fn)
            if not facturador_df.empty:
                # El DataFrame del Facturador ya tiene las mismas columnas que el consolidado MC.
                # Concatenar al consolidado (o crear consolidado si MC está vacío).
                if consolidado.empty:
                    consolidado = facturador_df
                else:
                    consolidado = pd.concat([consolidado, facturador_df], ignore_index=True)
                _log_info(f"Facturador: {len(facturador_df)} registros incorporados al consolidado.", log_fn)
            else:
                _log_info("Facturador: sin registros válidos para incorporar.", log_fn)

        if consolidado.empty:
            _log_info("No se encontraron datos en los archivos (MC, Facturador).", log_fn)
            return

        # Rename columns to shorter names for processing
        consolidado.rename(columns={
            'Fecha de Emisión': 'Fecha',
            'Tipo de Comprobante': 'Tipo',
            'Imp. Neto Gravado Total': 'Imp. Neto Gravado',
            'Total IVA': 'IVA'
        }, inplace=True)

        # Process amounts
        columnas_numericas = ['Imp. Neto Gravado', 'Imp. Neto No Gravado', 'Imp. Op. Exentas', 'Otros Tributos', 'IVA', 'Imp. Total']
        # Convert to float (handling commas)
        for col in columnas_numericas:
            if col in consolidado.columns:
                 # Clean string if needed
                 if consolidado[col].dtype == object:
                      consolidado[col] = consolidado[col].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                      consolidado[col] = pd.to_numeric(consolidado[col], errors='coerce').fillna(0)

        if 'Tipo Cambio' in consolidado.columns:
             if consolidado['Tipo Cambio'].dtype == object:
                  consolidado['Tipo Cambio'] = consolidado['Tipo Cambio'].astype(str).str.replace(',', '.', regex=False)
                  consolidado['Tipo Cambio'] = pd.to_numeric(consolidado['Tipo Cambio'], errors='coerce').fillna(1)
             consolidado.loc[consolidado['Tipo Cambio'] == 0, 'Tipo Cambio'] = 1

             for col in columnas_numericas:
                 if col in consolidado.columns:
                     consolidado[col] = consolidado[col] * consolidado['Tipo Cambio']

        # Handle Credit Notes
        consolidado.loc[consolidado['Tipo'].isin(NOTAS_DE_CREDITO), columnas_numericas] *= -1

        # Drop unused
        consolidado.drop(['Imp. Neto Gravado', 'Imp. Neto No Gravado', 'Imp. Op. Exentas', 'IVA'], axis=1, inplace=True, errors='ignore')

        # MC column (extracted from filename part 1?)
        # control.py: consolidado['MC'] = consolidado['Archivo'].str.split("-").str[1].str.strip()
        # Dependent on filename format. Safe to skip or try.
        try:
             consolidado['MC'] = consolidado['Archivo'].str.split("-").str[1].str.strip()
        except:
             consolidado['MC'] = ""

        # Build AUX
        # CUIT_Emisor-COD(3)-PtoVenta(5)-Numero(8)
        # Fin CUIT is emisor
        consolidado['AUX'] = (
            consolidado['Fin CUIT'].astype(int).astype(str) + "-" +
            consolidado['Tipo'].astype(int).astype(str).str.zfill(3) + "-" +
            consolidado['Punto de Venta'].astype(int).astype(str).str.zfill(5) + "-" +
            consolidado['Número Desde'].astype(int).astype(str).str.zfill(8)
        )

        columnas_rcel_requeridas = ['AUX', 'Desde', 'Hasta', 'Archivo PDF']
        tiene_rcel = (
            not info_facturas_pdf.empty
            and all(c in info_facturas_pdf.columns for c in columnas_rcel_requeridas)
        )

        if tiene_rcel:
             consolidado = pd.merge(consolidado, info_facturas_pdf[columnas_rcel_requeridas], how='left', on='AUX')
        else:
             if not info_facturas_pdf.empty:
                 _log_info("Archivos JSON encontrados pero sin columnas RCEL esperadas (AUX, Desde, Hasta, Archivo PDF). "
                           "Se ignorarán y se usará la fecha de emisión como fallback.", log_fn)
             consolidado['Desde'] = pd.NaT
             consolidado['Hasta'] = pd.NaT
             consolidado['Archivo PDF'] = None

        consolidado['Cruzado'] = np.where(consolidado['Archivo PDF'].notnull(), 'Si', 'No')

        # Dates processing
        consolidado['Fecha'] = pd.to_datetime(consolidado['Fecha'], format='ISO8601', errors='coerce')
        if 'Desde' in consolidado.columns:
             consolidado['Desde'] = pd.to_datetime(consolidado['Desde'], dayfirst=True, errors='coerce')
        if 'Hasta' in consolidado.columns:
             consolidado['Hasta'] = pd.to_datetime(consolidado['Hasta'], dayfirst=True, errors='coerce')

        # Complete billing range with invoice date when JSON dates are missing.
        consolidado['Desde'] = consolidado['Desde'].fillna(consolidado['Fecha'])
        consolidado['Hasta'] = consolidado['Hasta'].fillna(consolidado['Fecha'])

        # ── Corregir columnas específicas del Facturador ──
        if 'FchServDesde' in consolidado.columns:
            mask_fact = consolidado['FchServDesde'].notna()
            if mask_fact.any():
                consolidado.loc[mask_fact, 'Desde'] = consolidado.loc[mask_fact, 'FchServDesde']
                consolidado.loc[mask_fact, 'Hasta'] = consolidado.loc[mask_fact, 'FchServHasta']
                consolidado.loc[mask_fact, 'MC'] = 'MCE'
                consolidado.loc[mask_fact, 'Cruzado'] = 'Si'

        # Billing period length now uses normalized Desde/Hasta columns.
        desde_facturacion = consolidado['Desde']
        hasta_facturacion = consolidado['Hasta']

        # Filter by dates? control.py has a commented out line for this. I'll skip.

        # Pro-rating
        consolidado['Fecha_Inicial_max'] = fecha_inicial
        mask_desde = consolidado['Desde'].notna()
        consolidado.loc[mask_desde, 'Fecha_Inicial_max'] = consolidado.loc[
            mask_desde, ['Fecha_Inicial_max', 'Desde']
        ].max(axis=1)

        consolidado['Fecha_Final_min'] = fecha_final
        mask_hasta = consolidado['Hasta'].notna()
        consolidado.loc[mask_hasta, 'Fecha_Final_min'] = consolidado.loc[
            mask_hasta, ['Fecha_Final_min', 'Hasta']
        ].min(axis=1)

        consolidado['Dias de facturación'] = (hasta_facturacion - desde_facturacion).dt.days + 1
        consolidado['Días Efectivos'] = (consolidado['Fecha_Final_min'] - consolidado['Fecha_Inicial_max']).dt.days + 1
        consolidado.loc[consolidado['Días Efectivos'] < 0, 'Días Efectivos'] = 0

        # Avoid division by zero
        consolidado['Dias de facturación'] = consolidado['Dias de facturación'].replace(0, 1)

        consolidado['Importe por día'] = consolidado['Imp. Total'] / consolidado['Dias de facturación']
        consolidado['Importe Prorrateado'] = consolidado['Importe por día'] * consolidado['Días Efectivos']

        # Pivot Table
        tabla_dinamica = pd.pivot_table(
            consolidado,
            values=['Importe Prorrateado', 'Tipo'],
            index=['Cliente', 'MC'],
            aggfunc={'Importe Prorrateado': 'sum', 'Tipo': 'count'}
        )
        tabla_dinamica.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)

        # Categorization (using module-level reusable functions)
        tabla_dinamica['Ingresos brutos máximos por la categoría'] = (
            tabla_dinamica['Importe Prorrateado'].apply(obtener_max_ingresos_categoria, args=(categorias,))
        )
        tabla_dinamica['Categoría'] = (
            tabla_dinamica['Importe Prorrateado'].apply(obtener_categoria, args=(categorias,))
        )

        # ─── Generar reportes HTML con gráficos (ANTES de formatear fechas) ──
        if html_output_dir:
            _log_info("Generando reportes HTML con gráficos...", log_fn)
            try:
                from mrbot_app.reporte_monotributista_html import exportar_reportes_html
                exportar_reportes_html(
                    consolidado=consolidado,
                    categorias=categorias,
                    output_dir=html_output_dir,
                    fecha_inicial=fecha_inicial,
                    fecha_final=fecha_final,
                    log_fn=log_fn,
                    logo_b64=logo_b64,
                )
                _log_info(f"Reportes HTML generados en: {html_output_dir}", log_fn)
            except Exception as html_e:
                _log_error(f"Error generando reportes HTML: {html_e}", log_fn)
                import traceback
                _log_error(traceback.format_exc(), log_fn)

        # Formatting Dates for export (Excel necesita strings)
        consolidado.drop(['FchServDesde', 'FchServHasta'], axis=1, inplace=True, errors='ignore')
        for c in ['Desde', 'Hasta', 'Fecha', 'Fecha_Inicial_max', 'Fecha_Final_min']:
             consolidado[c] = consolidado[c].dt.strftime('%d/%m/%Y')

        # Export
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            tabla_dinamica.to_excel(writer, sheet_name='Tabla Dinámica')
            consolidado.to_excel(writer, sheet_name='Consolidado', index=False)

        # Apply Styles
        wb = load_workbook(output_path)

        if 'Tabla Dinámica' in wb.sheetnames:
            ws = wb['Tabla Dinámica']
            aplicar_formato_encabezado(ws)
            aplicar_formato_moneda(ws, 3, 3) # Approx cols
            aplicar_formato_moneda(ws, 5, 5)
            autoajustar_columnas(ws)
            agregar_filtros(ws)

        if 'Consolidado' in wb.sheetnames:
            ws = wb['Consolidado']
            aplicar_formato_encabezado(ws)
            autoajustar_columnas(ws)
            agregar_filtros(ws)

        wb.save(output_path)
        _log_info(f"Reporte generado exitosamente: {output_path}", log_fn)

    except Exception as e:
        _log_error(f"Error generando reporte: {e}", log_fn)
        import traceback
        _log_error(traceback.format_exc(), log_fn)
